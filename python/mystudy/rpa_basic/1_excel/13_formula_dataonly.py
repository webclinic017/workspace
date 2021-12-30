from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None이라고 표시 (계산되어지려면 엑셀을 수동으로 열어서 저장해주고 다시 실행시켜야 함)
for row in ws.values:
    for cell in row:
        print(cell)